#!/usr/bin/env python3
"""Build final Excel from all findings — OCR + visual inspection."""
import json, os, urllib.request, urllib.error, time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

OUT_DIR = '/Users/ty/Downloads/quantscience_output'

# ══════════════════════════════════════════════════════════════
# CURATED DATA — consolidated from OCR extraction + visual spot-checks
# ══════════════════════════════════════════════════════════════

papers = [
    {
        'title': 'A First Look at Financial Data Analysis Using ChatGPT-4o',
        'authors': 'Zifeng Feng, Bingxin Li, Feng Liu',
        'year': '2024',
        'id': 'SSRN-4849578',
        'url': 'https://papers.ssrn.com/sol3/papers.cfm?abstract_id=4849578',
        'tags': 'llm, data-analysis, finance',
        'caption': 'How to do financial data analysis using ChatGPT. A 54-page PDF.',
        'source_ids': '1f150a1610b576c05838b6af3d0992b7, 43cc68a186b98b3edc9af82e718fa4df, 98b21e9aa44b070d502c810b80e1a030',
        'confidence': 'high',
    },
    {
        'title': '151 Trading Strategies',
        'authors': 'Zura Kakushadze, Juan Andrés Serur',
        'year': '2018',
        'id': 'SSRN-3247865',
        'url': 'https://papers.ssrn.com/sol3/papers.cfm?abstract_id=3247865',
        'tags': 'trading-strategies, systematic',
        'caption': '151 Trading Strategies. A 361-page PDF.',
        'source_ids': '',
        'confidence': 'high',
    },
    {
        'title': 'Pairs Trading: Performance of a Relative Value Arbitrage Rule',
        'authors': 'Evan Gatev, William Goetzmann, K. Geert Rouwenhorst',
        'year': '2006',
        'id': '',
        'url': '',
        'tags': 'pairs-trading, stat-arb, relative-value',
        'caption': 'A 47-page PDF explains the secrets of Pairs Trading used by hedge funds.',
        'source_ids': '97a70eb5a5709e1fecdc2b86ba8ffa65, 3a29f4dcf992de565200d0a9eae0620a, a79f0b6d452d175c7ece9512c801ad1e',
        'confidence': 'high',
    },
    {
        'title': 'Time Series Momentum',
        'authors': 'Tobias J. Moskowitz, Yao Hua Ooi, Lasse Heje Pedersen',
        'year': '2012',
        'id': '',
        'url': '',
        'tags': 'momentum, time-series, asset-classes',
        'caption': 'Time Series Momentum. A 23-page PDF. Here are the best parts.',
        'source_ids': 'aa25deda68538a6f89fc2ab81f54fc92, de6e13a1d0ad8295f0400463eb1c1907',
        'confidence': 'high',
    },
    {
        'title': 'Financial Statement Analysis with Large Language Models',
        'authors': 'Alex G. Kim, Maximilian Muhn, Valeri V. Nikolaev',
        'year': '2024',
        'id': '',
        'url': '',
        'tags': 'llm, financial-statements, gpt4',
        'caption': 'Financial Statement Analysis with Large Language Models (LLMs). A 54-page PDF.',
        'source_ids': '819325f4247c8c38e16fca57739ab7a7, 065078cfffa19df6c2ce693a6ab1541b, cc458f2bde34bff5d7aeb839bbeafa6e, 94f0da6bc0049e6a2e901fcbd1574ffc, 083fb7d8d6978328e1cbb55ac72c7b3e, f19ddc485ecb4729f6c6cbd539447d2f',
        'confidence': 'high',
    },
    {
        'title': 'Breaking the Trend: How to Avoid Cherry-Picked Signals',
        'authors': 'Sebastien Valeyre',
        'year': '2025',
        'id': 'arxiv-2504.10914',
        'url': 'https://arxiv.org/abs/2504.10914',
        'tags': 'trend-following, signal-selection, CTA',
        'caption': 'BREAKING: A new 33-page PDF demystifying how hedge funds create bias-free signals.',
        'source_ids': '9218883f9aa56155ac6838b333255836, aa63b67c223bb7f89bb780c870d82d13, dec022785a053d410b7e8e03fab50173',
        'confidence': 'high',
    },
    {
        'title': 'The Alchemy of Multibagger Stocks',
        'authors': '',
        'year': '',
        'id': '',
        'url': '',
        'tags': 'stock-picking, factor-investing, outperformance',
        'caption': 'This paper analyzed 464 stocks that 10X-ed over a 24-year period. 41-page PDF.',
        'source_ids': '5957d5a97a6430c2cf15cc92129b7849, 2c264f619c1f789e3fd4be5e2e273959, 1fffaf750567e747c542b7e41a5886d9, 62009a6873088822dee8f4c6950ee687, 697462b940051b9e8c847e9c0f94f6d9',
        'confidence': 'medium',
    },
    {
        'title': 'Using Mathematics to Make Money',
        'authors': 'Jim Simons',
        'year': '',
        'id': '',
        'url': '',
        'tags': 'quant-legends, mathematics, trading',
        'caption': '*Using Mathematics to Make Money* by legend Jim Simons. 8-page PDF free for download.',
        'source_ids': '1a047c680590d877e33ef49c3048680d, 4cc9cb3caa39cccf5d2d57316c7ef324, 01784eb791c8d180b7f69b9dbec41cef, 202d75c943789474e8a74226b8c89392',
        'confidence': 'high',
    },
    {
        'title': 'Bridging Language Models and Financial Analysis',
        'authors': 'Alejandro Lopez-Lira, Jihoon Kwon, Sangwoon Yoon, Jy-yong Sohn, Chanyeol Choi',
        'year': '2025',
        'id': '',
        'url': '',
        'tags': 'llm, finance, survey',
        'caption': 'BREAKING: Large Language Models and Financial Analysis. A new 28-page paper.',
        'source_ids': 'c66168bd4d659e8bae33d4d5f54fe978',
        'confidence': 'high',
    },
    {
        'title': 'DeepSeek-R1: Incentivizing Reasoning Capability in LLMs via Reinforcement Learning',
        'authors': 'DeepSeek-AI',
        'year': '2025',
        'id': 'arxiv-2501.12948',
        'url': 'https://arxiv.org/abs/2501.12948',
        'tags': 'llm, reasoning, reinforcement-learning',
        'caption': 'Chinese Quant Fund Turned AI Pioneer. The birth of DeepSeek.',
        'source_ids': 'cbd3fa87a00230db461a1b3d3ca14013',
        'confidence': 'high',
    },
    {
        'title': 'Clustering Market Regimes Using the Wasserstein Distance',
        'authors': '',
        'year': '',
        'id': '',
        'url': '',
        'tags': 'market-regimes, clustering, optimal-transport',
        'caption': 'The secret hedge funds use to detect market regimes. A 37-page PDF.',
        'source_ids': '5575b8784b2b78a19e270e4a76fb1ac2',
        'confidence': 'medium',
    },
    {
        'title': 'Fundamentals of Building Autonomous LLM Agents',
        'authors': 'Victor de Lamo Castrillo, Habtom Kahsay Gidey, Alexander Lenz, Alois Knoll',
        'year': '2025',
        'id': '',
        'url': '',
        'tags': 'llm-agents, architecture, agentic-ai',
        'caption': 'The Fundamentals Of Building Autonomous LLM Agents. A 38-page PDF.',
        'source_ids': 'a364dd76e57fb13e3f62d1d72c0e5a7a, 8ff8da47ddd868d82fa2e94288e8ddd4',
        'confidence': 'high',
    },
    {
        'title': 'Context Engineering 2.0: The Context of Context Engineering',
        'authors': 'Qishuo Hua, Lyumanshan Ye, Dayuan Fu, Yang Xiao, Xiaojie Cai, Yunze Wu, Jifan Lin, Junfei Wang, Pengfei Liu',
        'year': '2026',
        'id': '',
        'url': '',
        'tags': 'llm, context-engineering, human-ai-interaction',
        'caption': 'RIP Prompt Engineering. Enter Context Engineering 2.0. A 28-page PDF.',
        'source_ids': '54267c86c7e3119cbf2af7ec4926b66f, 9389c7ebce07650fae8175ce743ca9a2, b444cdbcdc7c6f0ea307f937e8883533',
        'confidence': 'high',
    },
    {
        'title': 'Teaching LLMs to Plan: Logical Chain-of-Thought Instruction Tuning for Symbolic Planning',
        'authors': 'Pulkit Verma, Ngoc La, Anthony Favier, Swaroop Mishra, Julie A. Shah',
        'year': '2025',
        'id': 'arxiv-2509.13351',
        'url': 'https://arxiv.org/abs/2509.13351',
        'tags': 'llm, planning, chain-of-thought',
        'caption': 'BREAKING: MIT researchers discover how to enable LLMs to do real logical reasoning.',
        'source_ids': '482e687ebc9cea17e9352e3c28994702',
        'confidence': 'high',
    },
    {
        'title': 'A Practical Guide to Quantitative Volatility Trading',
        'authors': 'Daniel Bloch',
        'year': '2016',
        'id': 'SSRN-2715517',
        'url': 'https://papers.ssrn.com/sol3/papers.cfm?abstract_id=2715517',
        'tags': 'volatility, options, quantitative-trading',
        'caption': 'Jane Street, AQR, Ren Tech... All use volatility. A 327-page PDF.',
        'source_ids': 'f47b912958a34983cf436d7cc01508a1',
        'confidence': 'high',
    },
    {
        'title': 'Financial Machine Learning (S&P Global Market Intelligence)',
        'authors': '',
        'year': '',
        'id': '',
        'url': '',
        'tags': 'ml-for-finance, algorithmic-trading',
        'caption': '159 page PDF download. The best examples of how machine learning is used in finance.',
        'source_ids': 'c43c36cf6cf60d075f76e1f80a5dd6ae, 40a4097a940c78f644160568c5406fb3, e08e4d95bf3608d6d243b25f6f6906c6, f203a9ac95667ac3be89823f1f7697e2',
        'confidence': 'medium',
    },
    {
        'title': 'Machine learning techniques and data for stock market forecasting: A literature review',
        'authors': '',
        'year': '',
        'id': '',
        'url': '',
        'tags': 'ml-for-finance, stock-forecasting, survey',
        'caption': 'Machine learning techniques and data for stock market forecasting. A free 41-page PDF.',
        'source_ids': '4b34c2f20c76a2c1b383e05daa44b5a4, deee5b7f80d78b9286de6a40d332637c',
        'confidence': 'medium',
    },
    {
        'title': 'Empirical Asset Pricing via Machine Learning (ranking 94 factors)',
        'authors': 'Shihao Gu, Bryan Kelly, Dacheng Xiu',
        'year': '2020',
        'id': '',
        'url': '',
        'tags': 'factor-investing, ml-for-finance, asset-pricing',
        'caption': 'Ranking 94 factors in algorithmic trading with Machine Learning. A 51-page paper.',
        'source_ids': 'cd806b211b2de90cfc0553d8a70f441b',
        'confidence': 'medium',
    },
    {
        'title': 'Agents Companion (Google Whitepaper)',
        'authors': 'Google',
        'year': '2026',
        'id': '',
        'url': '',
        'tags': 'llm-agents, mlops, whitepaper',
        'caption': 'AGENTS COMPANION: A NEW WHITEPAPER BY GOOGLE. 76-page PDF.',
        'source_ids': 'b34accc569a4d402e22653b0d9a7a2e4',
        'confidence': 'medium',
    },
    {
        'title': 'Building Cross-Sectional Systematic Strategies',
        'authors': '',
        'year': '',
        'id': '',
        'url': '',
        'tags': 'systematic-trading, cross-sectional',
        'caption': 'Building Cross-Sectional Systematic Strategies.',
        'source_ids': '',
        'confidence': 'low',
    },
    {
        'title': 'Python for Finance: Analyze Big Financial Data',
        'authors': 'Yves Hilpisch',
        'year': '',
        'id': '',
        'url': '',
        'tags': 'python, finance, book',
        'caption': 'Python for Finance: A beginner-friendly intro to Python for financial analysis.',
        'source_ids': 'd76fe2600c6b971191bca0afaa658e89',
        'confidence': 'high',
    },
    {
        'title': 'TradingAgents: Multi-Agents LLM Financial Trading Framework',
        'authors': '',
        'year': '2025',
        'id': '',
        'url': 'https://github.com/TauricResearch/TradingAgents',
        'tags': 'llm-agents, trading, multi-agent',
        'caption': 'BREAKING: A new open-source multi-agent LLM trading framework. TradingAgents.',
        'source_ids': '5ab5a76e3609e67c5ac8ad7304a61cee, 507d322538e35675e6643f3ed5a37401',
        'confidence': 'high',
    },
]

repos = [
    {
        'name': 'OpenBBTerminal',
        'github_url': 'https://github.com/OpenBB-finance/OpenBBTerminal',
        'language': 'Python',
        'category': 'data, terminal',
        'description': 'Investment research terminal — open-source Bloomberg alternative.',
        'caption': 'BREAKING: New Python Library for Finance Analysis with AI Agents.',
        'source_ids': 'multiple (72+ posts)',
        'confidence': 'high',
    },
    {
        'name': 'openbb-agents',
        'github_url': 'https://github.com/OpenBB-finance/openbb-agents',
        'language': 'Python',
        'category': 'llm-agent, data',
        'description': 'LLM agents that use OpenBB Platform for autonomous financial research.',
        'caption': 'New Python Library for Finance Analysis with AI Agents. pip install openbb-agents.',
        'source_ids': '4e9b52e406486dbc04836066c694e7ae',
        'confidence': 'high',
    },
    {
        'name': 'python-training',
        'github_url': 'https://github.com/jpmorganchase/python-training',
        'language': 'Python',
        'category': 'education',
        'description': "JP Morgan's Python training materials for finance.",
        'caption': "JP Morgan's Python training. Available 100% for free.",
        'source_ids': '6b043d2aebd47afc39a0f26d9898d520, f59652847559fb8690fc2a8054506308',
        'confidence': 'high',
    },
    {
        'name': 'qlib',
        'github_url': 'https://github.com/microsoft/qlib',
        'language': 'Python',
        'category': 'ml, trading-framework',
        'description': 'Microsoft AI-oriented quantitative investment platform.',
        'caption': 'BREAKING: Microsoft open-sourced an AI Quant investment platform in Python.',
        'source_ids': '804b54c578961d23aca82ad6b0394ac2, 82923a071bcfd150066e5c4b927fa11c, 580e4b183c2ca2f8e6efac7e61ad4121, eb5da55c95951d4e9dd0b5f39fbcfa6f, c2564d0de0c1efd5dd02a4f36e42ee0e',
        'confidence': 'high',
    },
    {
        'name': 'RD-Agent',
        'github_url': 'https://github.com/microsoft/RD-Agent',
        'language': 'Python',
        'category': 'ml, automated-research',
        'description': 'Automated quant R&D pipeline by Microsoft.',
        'caption': 'Microsoft open-sourced an AI Quant investment platform.',
        'source_ids': '804b54c578961d23aca82ad6b0394ac2',
        'confidence': 'high',
    },
    {
        'name': 'gs-quant',
        'github_url': 'https://github.com/goldmansachs/gs-quant',
        'language': 'Python',
        'category': 'quant-toolkit',
        'description': 'Goldman Sachs Python toolkit for quantitative finance.',
        'caption': 'Introducing Goldman Sachs GS-Quant. A Python quant toolkit made by Goldman Sachs.',
        'source_ids': 'a1a062b604da75df98d70d56f12e9a3f',
        'confidence': 'high',
    },
    {
        'name': 'Riskfolio-Lib',
        'github_url': 'https://github.com/dcajasn/Riskfolio-Lib',
        'language': 'Python',
        'category': 'portfolio-optimization',
        'description': 'Portfolio optimization and quantitative strategic asset allocation.',
        'caption': 'Python library for portfolio optimization.',
        'source_ids': 'multiple (16+ posts)',
        'confidence': 'high',
    },
    {
        'name': 'tensortrade',
        'github_url': 'https://github.com/tensortrade-org/tensortrade',
        'language': 'Python',
        'category': 'trading-framework, reinforcement-learning',
        'description': 'Open-source framework for trading using reinforcement learning.',
        'caption': 'BREAKING: A new Python library for algorithmic trading. Introducing TensorTrade.',
        'source_ids': '03aba6ec8a5218e8166ac49d06ffa173, 347ca375c2440ecf1ec9129d0b12e4dd, 4c8775c57c3addd576a911692ae0edce',
        'confidence': 'high',
    },
    {
        'name': 'vectorbt',
        'github_url': 'https://github.com/polakowo/vectorbt',
        'language': 'Python',
        'category': 'backtesting',
        'description': 'Vectorized backtesting and trading framework.',
        'caption': 'Python library for vectorized backtesting.',
        'source_ids': 'multiple (13+ posts)',
        'confidence': 'high',
    },
    {
        'name': 'freqtrade',
        'github_url': 'https://github.com/freqtrade/freqtrade',
        'language': 'Python',
        'category': 'trading-framework, crypto',
        'description': 'Free, open-source crypto trading bot.',
        'caption': 'Open-source crypto trading framework.',
        'source_ids': 'multiple (3+ posts)',
        'confidence': 'high',
    },
    {
        'name': 'nautilus_trader',
        'github_url': 'https://github.com/nautechsystems/nautilus_trader',
        'language': 'Python/Rust',
        'category': 'trading-framework',
        'description': 'High-performance algorithmic trading platform.',
        'caption': "BREAKING: Python's Newest Algorithmic Trading Tool. Introducing Nautilus Trader.",
        'source_ids': '5bef5355323507926f50a6a4b3e7a68f',
        'confidence': 'high',
    },
    {
        'name': 'FinRL',
        'github_url': 'https://github.com/AI4Finance-Foundation/FinRL',
        'language': 'Python',
        'category': 'reinforcement-learning, trading',
        'description': 'Deep reinforcement learning framework for quantitative finance.',
        'caption': 'FinRL: deep reinforcement learning for finance.',
        'source_ids': 'multiple (2+ posts)',
        'confidence': 'high',
    },
    {
        'name': 'QuantLib',
        'github_url': 'https://github.com/lballabio/QuantLib',
        'language': 'C++',
        'category': 'derivatives, pricing',
        'description': 'Quantitative finance library for modeling, pricing, and risk management.',
        'caption': 'QuantLib for quantitative finance.',
        'source_ids': 'multiple (3+ posts)',
        'confidence': 'high',
    },
    {
        'name': 'pytimetk',
        'github_url': 'https://github.com/business-science/pytimetk',
        'language': 'Python',
        'category': 'time-series, data',
        'description': 'Time series analysis toolkit for Python.',
        'caption': 'This Python library is wild for financial analysis. It\'s called Pytimetk.',
        'source_ids': '9888fe50e9296a02e0bf783fdf543d01',
        'confidence': 'high',
    },
    {
        'name': 'ib_insync',
        'github_url': 'https://github.com/erdewit/ib_insync',
        'language': 'Python',
        'category': 'broker-api',
        'description': 'Third-party Interactive Brokers API wrapper.',
        'caption': 'Interactive Brokers API connection with Python.',
        'source_ids': '',
        'confidence': 'high',
    },
    {
        'name': 'quantstats',
        'github_url': 'https://github.com/ranaroussi/quantstats',
        'language': 'Python',
        'category': 'analytics, reporting',
        'description': 'Portfolio analytics for quants — performance and risk metrics.',
        'caption': 'QuantStats for portfolio analytics.',
        'source_ids': 'multiple (2+ posts)',
        'confidence': 'high',
    },
    {
        'name': 'yfinance',
        'github_url': 'https://github.com/ranaroussi/yfinance',
        'language': 'Python',
        'category': 'data',
        'description': 'Download market data from Yahoo Finance.',
        'caption': 'yfinance for market data.',
        'source_ids': '',
        'confidence': 'high',
    },
    {
        'name': 'alphalens',
        'github_url': 'https://github.com/quantopian/alphalens',
        'language': 'Python',
        'category': 'factor-analysis',
        'description': 'Performance analysis of predictive (alpha) stock factors.',
        'caption': 'Alphalens for alpha factor analysis.',
        'source_ids': 'multiple (2+ posts)',
        'confidence': 'high',
    },
    {
        'name': 'dspy',
        'github_url': 'https://github.com/stanfordnlp/dspy',
        'language': 'Python',
        'category': 'llm, prompt-engineering',
        'description': 'Stanford framework for programming with foundation models.',
        'caption': 'DSPy for programming LLMs.',
        'source_ids': 'multiple (2+ posts)',
        'confidence': 'high',
    },
    {
        'name': 'docling',
        'github_url': 'https://github.com/DS4SD/docling',
        'language': 'Python',
        'category': 'document-processing',
        'description': 'Document understanding and conversion toolkit.',
        'caption': 'Docling for document processing.',
        'source_ids': 'multiple (3+ posts)',
        'confidence': 'high',
    },
    {
        'name': 'julius',
        'github_url': 'https://github.com/julius-ai/julius',
        'language': 'Python',
        'category': 'llm-agent, data-analysis',
        'description': 'AI data analyst agent.',
        'caption': 'Julius AI for data analysis.',
        'source_ids': 'multiple (2+ posts)',
        'confidence': 'medium',
    },
    {
        'name': 'Ziplime',
        'github_url': '',
        'language': 'Python',
        'category': 'trading-framework, ai',
        'description': 'AI-powered algorithmic trading library (not Zipline).',
        'caption': "BREAKING: New Python Library for Algorithmic Trading with AI. It's called Ziplime (*not* Zipline).",
        'source_ids': '22d3b9a15ec88877a694d37e7dbb8f04',
        'confidence': 'medium',
    },
    {
        'name': 'QF-Lib',
        'github_url': 'https://github.com/quarkfin/qf-lib',
        'language': 'Python',
        'category': 'quant-toolkit, backtesting',
        'description': 'Python library for quant finance — backtesting, analysis, reporting.',
        'caption': 'BREAKING: Introducing QF-Lib. A new Python library for Quant Finance.',
        'source_ids': 'cdd17534761b89e7a9f8761116ae816e',
        'confidence': 'medium',
    },
    {
        'name': 'AlphaPy',
        'github_url': 'https://github.com/ScottFreeLLC/AlphaPy',
        'language': 'Python',
        'category': 'ml, trading-framework',
        'description': 'Machine learning framework for algorithmic trading and sports betting.',
        'caption': "BREAKING: I just stumbled upon this ML Python library for Algorithmic Trading. It's called AlphaPy.",
        'source_ids': 'e1fff37b822782c3140d2ce877ab7c7a',
        'confidence': 'medium',
    },
    {
        'name': 'AI Hedge Fund (virattt)',
        'github_url': 'https://github.com/virattt/ai-hedge-fund',
        'language': 'Python',
        'category': 'llm-agent, trading',
        'description': 'Open-source AI hedge fund — multi-agent architecture for trading.',
        'caption': 'This guy made an AI Hedge Fund... Then open-sourced it for everyone to use.',
        'source_ids': '33897e17e54f0bd401058808f8918474',
        'confidence': 'medium',
    },
    {
        'name': 'AI Financial Agent',
        'github_url': 'https://github.com/virattt/ai-financial-agent',
        'language': 'Python',
        'category': 'llm-agent, investment-research',
        'description': 'Proof-of-concept AI financial agent for investment research.',
        'caption': 'This guy made an AI financial agent for investment research. Then gave away the code for free.',
        'source_ids': '21b9325ee701ce3717bb6fb10da63e12, 316342604e00869468fe9349c5ddabb4',
        'confidence': 'low',
    },
    {
        'name': 'Stock Research Agent',
        'github_url': '',
        'language': 'Python',
        'category': 'llm-agent, investment-research',
        'description': 'AI platform for financial research. Built on LangGraph + LangSmith. Version 3.',
        'caption': 'Stock Research Agent made in Python. 100% open source.',
        'source_ids': '2dde835108a3de2fed8cafce6d33fc75',
        'confidence': 'low',
    },
    {
        'name': 'TradingAgents',
        'github_url': 'https://github.com/TauricResearch/TradingAgents',
        'language': 'Python',
        'category': 'llm-agent, trading',
        'description': 'Multi-agent LLM financial trading framework.',
        'caption': 'BREAKING: A new open-source multi-agent LLM trading framework. TradingAgents.',
        'source_ids': '5ab5a76e3609e67c5ac8ad7304a61cee, 507d322538e35675e6643f3ed5a37401',
        'confidence': 'medium',
    },
    {
        'name': 'Nanobot',
        'github_url': 'https://github.com/nanobot-ai/nanobot',
        'language': 'Python',
        'category': 'llm-agent, mcp',
        'description': 'Build MCP AI Agents with reasoning, system prompts, and tool orchestration.',
        'caption': 'Introducing Nanobot. Build MCP AI Agents.',
        'source_ids': '646714a74aa4d8e3166defe487285600',
        'confidence': 'medium',
    },
    {
        'name': 'stable-diffusion',
        'github_url': 'https://github.com/CompVis/stable-diffusion',
        'language': 'Python',
        'category': 'generative-ai',
        'description': 'Latent text-to-image diffusion model.',
        'caption': 'Stable Diffusion mentioned in AI context.',
        'source_ids': '',
        'confidence': 'medium',
    },
    {
        'name': 'claude-cookbooks',
        'github_url': 'https://github.com/anthropics/claude-cookbooks',
        'language': 'Python',
        'category': 'llm, cookbook',
        'description': "Anthropic's Claude API usage examples and recipes.",
        'caption': 'Claude cookbooks for AI development.',
        'source_ids': '',
        'confidence': 'medium',
    },
    {
        'name': 'openai-cookbook',
        'github_url': 'https://github.com/openai/openai-cookbook',
        'language': 'Python',
        'category': 'llm, cookbook',
        'description': "OpenAI's API usage examples and best practices.",
        'caption': 'OpenAI cookbook for AI development.',
        'source_ids': '',
        'confidence': 'medium',
    },
    {
        'name': 'AI-For-Beginners',
        'github_url': 'https://github.com/microsoft/AI-For-Beginners',
        'language': 'Jupyter Notebook',
        'category': 'education',
        'description': 'Microsoft 12-week AI curriculum.',
        'caption': 'Microsoft AI for beginners curriculum.',
        'source_ids': '',
        'confidence': 'medium',
    },
    {
        'name': 'Data-Science-For-Beginners',
        'github_url': 'https://github.com/microsoft/Data-Science-For-Beginners',
        'language': 'Jupyter Notebook',
        'category': 'education',
        'description': 'Microsoft 10-week data science curriculum.',
        'caption': 'Microsoft Data Science for beginners.',
        'source_ids': '',
        'confidence': 'medium',
    },
    {
        'name': 'Python-100-Days',
        'github_url': 'https://github.com/jackfrued/Python-100-Days',
        'language': 'Python',
        'category': 'education',
        'description': 'Python in 100 days — from novice to master.',
        'caption': 'Python 100 Days learning resource.',
        'source_ids': '',
        'confidence': 'medium',
    },
    {
        'name': 'PythonDataScienceHandbook',
        'github_url': 'https://github.com/jakevdp/PythonDataScienceHandbook',
        'language': 'Jupyter Notebook',
        'category': 'education, data-science',
        'description': 'Python Data Science Handbook by Jake VanderPlas.',
        'caption': 'Python Data Science Handbook.',
        'source_ids': '',
        'confidence': 'medium',
    },
    {
        'name': 'CLIP',
        'github_url': 'https://github.com/openai/CLIP',
        'language': 'Python',
        'category': 'ml, vision-language',
        'description': 'Contrastive Language-Image Pre-training by OpenAI.',
        'caption': 'CLIP by OpenAI.',
        'source_ids': '',
        'confidence': 'medium',
    },
    {
        'name': 'segment-anything',
        'github_url': 'https://github.com/facebookresearch/segment-anything',
        'language': 'Python',
        'category': 'ml, computer-vision',
        'description': 'Segment Anything Model (SAM) by Meta.',
        'caption': 'Segment Anything by Meta.',
        'source_ids': '',
        'confidence': 'medium',
    },
    {
        'name': 'google-research',
        'github_url': 'https://github.com/google-research/google-research',
        'language': 'Python',
        'category': 'ml, research',
        'description': 'Google Research code releases.',
        'caption': 'Google Research repository.',
        'source_ids': '',
        'confidence': 'medium',
    },
]


# ══════════════════════════════════════════════════════════════
# VERIFY repos via GitHub API
# ══════════════════════════════════════════════════════════════
def check_github_repo(url):
    """Check if a GitHub repo exists and get description."""
    if not url:
        return None
    parts = url.replace('https://github.com/', '').split('/')
    if len(parts) < 2:
        return None
    owner, repo = parts[0], parts[1]
    api_url = f'https://api.github.com/repos/{owner}/{repo}'
    try:
        req = urllib.request.Request(api_url, headers={'User-Agent': 'quantscience-scraper'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode())
            return {
                'description': data.get('description', ''),
                'language': data.get('language', ''),
                'stars': data.get('stargazers_count', 0),
                'url': data.get('html_url', url),
            }
    except (urllib.error.HTTPError, urllib.error.URLError, Exception) as e:
        print(f"  ! Could not verify {url}: {e}")
        return None

print("Verifying GitHub repos...")
for repo in repos:
    if repo['github_url']:
        info = check_github_repo(repo['github_url'])
        if info:
            if info['description'] and not repo['description']:
                repo['description'] = info['description']
            if info['language']:
                repo['language'] = info['language']
            if repo['confidence'] != 'high':
                repo['confidence'] = 'high'
            print(f"  OK: {repo['name']} ({info['stars']} stars)")
        else:
            if repo['confidence'] == 'high':
                repo['confidence'] = 'medium'
            print(f"  ?? {repo['name']} — could not verify")
        time.sleep(0.3)  # rate limit

# ══════════════════════════════════════════════════════════════
# BUILD EXCEL
# ══════════════════════════════════════════════════════════════
print("\nBuilding Excel...")
wb = Workbook()

# ── Papers sheet ──
ws_papers = wb.active
ws_papers.title = 'Papers'
paper_headers = ['title', 'authors', 'year', 'arxiv_or_ssrn_id', 'url', 'topic_tags', 'caption_snippet', 'source_post_id', 'confidence']
ws_papers.append(paper_headers)

for p in papers:
    ws_papers.append([
        p['title'],
        p['authors'],
        p['year'],
        p['id'],
        p['url'],
        p['tags'],
        p['caption'],
        p['source_ids'],
        p['confidence'],
    ])

# ── Repos sheet ──
ws_repos = wb.create_sheet('Repos')
repo_headers = ['name', 'github_url', 'language', 'category', 'description', 'caption_snippet', 'source_post_id', 'confidence']
ws_repos.append(repo_headers)

for r in repos:
    ws_repos.append([
        r['name'],
        r['github_url'],
        r['language'],
        r['category'],
        r['description'],
        r['caption'],
        r['source_ids'],
        r['confidence'],
    ])

# ── Formatting ──
col_widths = {
    'Papers': {'A': 60, 'B': 45, 'C': 8, 'D': 20, 'E': 55, 'F': 35, 'G': 80, 'H': 40, 'I': 12},
    'Repos': {'A': 25, 'B': 50, 'C': 12, 'D': 25, 'E': 60, 'F': 80, 'G': 40, 'H': 12},
}

for sheet_name, widths in col_widths.items():
    ws = wb[sheet_name]
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    # Bold + freeze header
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = 'A2'
    # Wrap text for caption columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=False)

xlsx_path = os.path.join(OUT_DIR, 'quantscience_findings.xlsx')
wb.save(xlsx_path)
print(f"Saved: {xlsx_path}")
print(f"  Papers: {len(papers)}")
print(f"  Repos: {len(repos)}")

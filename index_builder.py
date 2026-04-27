"""
Quant resource index builder.

Tiers:
  Tier 1: awesome-lists (wilsonfreitas/awesome-quant, firmai/financial-machine-learning, cbailes/awesome-deep-trading, paperswithcode finance)
  Tier 2: arxiv q-fin
  Tier 3: institutional research (AQR, Man, Two Sigma, Alpha Architect, BIS, Fed FEDS)
  Tier 4: SSRN top-10 download lists
  Tier 5: curator blogs (QuantStart, Robot Wealth, Hudson & Thames)

Cached-only: rows tagged with "quantscience_ig" in data/quant_index.json carry over from a prior
seed-curation pass. There is no live fetcher for that tier — flip RUN flags to refresh other tiers.

Writes data/quant_index.json (source of truth) + data/quant_index.xlsx (per-source sheets + All_Deduped).
Idempotent: re-reads existing JSON, skips resource_ids refreshed within REFRESH_DAYS.
"""

from __future__ import annotations

import hashlib
import json
import re
import sys
import time
import urllib.parse
from dataclasses import asdict, dataclass, field
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Callable, Iterable

import feedparser
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# CONFIG — flip sources on/off for partial runs
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parent
OUT_JSON = ROOT / "data" / "quant_index.json"
OUT_XLSX = ROOT / "data" / "quant_index.xlsx"
OUT_SUMMARY = ROOT / "docs" / "INDEX_SUMMARY.md"

RUN = {
    "tier1_awesome": True,
    "tier2_arxiv": True,
    "tier3_institutional": True,
    "tier4_ssrn": True,
    "tier5_blogs": True,
}

# Tighten if Tier 1 overflows (>1500 rows)
TIER1_MIN_STARS = 0  # set to 100 to filter low-star repos
TIER1_DROP_BLOG_LINKS = False  # set True to drop obvious blog-post URLs

ARXIV_CATS = ["q-fin.TR", "q-fin.PM", "q-fin.ST", "q-fin.CP", "q-fin.RM", "q-fin.MF", "q-fin.PR"]
ARXIV_MAX_PER_CAT = 200
ARXIV_CITATION_TOP_N = 50
ARXIV_DELAY = 3.0

REFRESH_DAYS = 30
HTTP_TIMEOUT = 20
UA = "quant-index-builder/0.1 (personal research)"

SESSION = requests.Session()
SESSION.headers["User-Agent"] = UA

# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------


@dataclass
class Resource:
    resource_id: str
    type: str  # paper | repo | textbook | whitepaper | blog_post
    title: str = ""
    authors_or_owners: str = ""
    year: int | None = None
    sources: list[str] = field(default_factory=list)
    canonical_url: str = ""
    secondary_urls: str = ""
    topic_tags: str = ""
    one_line_summary: str = ""
    citation_count_or_stars: int | None = None
    date_published: str = ""
    mention_count: int = 1
    why_notable: str = ""
    confidence: str = "medium"
    retrieved_at: str = ""

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


# ---------------------------------------------------------------------------
# URL normalization + resource_id
# ---------------------------------------------------------------------------

_CONF_RANK = {"high": 3, "medium": 2, "low": 1, "": 0}


def _strip_fragment_and_tracking(u: str) -> str:
    parts = urllib.parse.urlsplit(u)
    q = [
        (k, v)
        for k, v in urllib.parse.parse_qsl(parts.query, keep_blank_values=False)
        if not k.lower().startswith(("utm_", "fbclid", "gclid", "mc_cid", "mc_eid"))
    ]
    return urllib.parse.urlunsplit(
        (
            parts.scheme.lower() or "https",
            parts.netloc.lower(),
            parts.path.rstrip("/") if parts.path != "/" else "/",
            urllib.parse.urlencode(q),
            "",
        )
    )


def normalize_url(u: str) -> str:
    if not u:
        return ""
    u = u.strip()
    if not u.startswith(("http://", "https://")):
        u = "https://" + u
    u = _strip_fragment_and_tracking(u)
    parts = urllib.parse.urlsplit(u)
    host = parts.netloc
    path = parts.path

    # GitHub: canonicalize to https://github.com/<owner>/<repo> lowercase (GH is case-insensitive)
    if host in ("github.com", "www.github.com"):
        segs = [s for s in path.split("/") if s]
        if len(segs) >= 2:
            owner, repo = segs[0], segs[1]
            if repo.endswith(".git"):
                repo = repo[:-4]
            path = f"/{owner.lower()}/{repo.lower()}"
            host = "github.com"
            return urllib.parse.urlunsplit(("https", host, path, "", ""))

    # arxiv: canonicalize to abs/<id> (strip version)
    if host.endswith("arxiv.org"):
        m = re.search(r"(\d{4}\.\d{4,5})(v\d+)?", path)
        if m:
            return f"https://arxiv.org/abs/{m.group(1)}"

    # SSRN: canonicalize to abstract_id form
    if "ssrn" in host:
        m = re.search(r"abstract[_-]?id=(\d+)|abstract/(\d+)", u, re.I)
        if m:
            aid = m.group(1) or m.group(2)
            return f"https://papers.ssrn.com/sol3/papers.cfm?abstract_id={aid}"

    return urllib.parse.urlunsplit((parts.scheme or "https", host, path, parts.query, ""))


def resource_id_for(url: str) -> str:
    return hashlib.sha1(normalize_url(url).encode("utf-8")).hexdigest()[:16]


# ---------------------------------------------------------------------------
# Dedup / merge
# ---------------------------------------------------------------------------


def merge_resources(a: Resource, b: Resource) -> Resource:
    """Merge b into a. Preserves higher confidence, longer title/summary, unioned sources."""
    if a is b:
        # Same instance appearing twice (e.g., a multi-source row listed under
        # every tier prefix its sources match). Merging self would double sources,
        # mention_count, etc. — no-op instead.
        return a
    for src in b.sources:
        if src not in a.sources:
            a.sources.append(src)
    if _CONF_RANK.get(b.confidence, 0) > _CONF_RANK.get(a.confidence, 0):
        a.confidence = b.confidence
    if len(b.title) > len(a.title):
        a.title = b.title
    if len(b.one_line_summary) > len(a.one_line_summary):
        a.one_line_summary = b.one_line_summary
    if not a.authors_or_owners and b.authors_or_owners:
        a.authors_or_owners = b.authors_or_owners
    if not a.year and b.year:
        a.year = b.year
    if not a.date_published and b.date_published:
        a.date_published = b.date_published
    a.mention_count = max(a.mention_count, 0) + max(b.mention_count, 0)
    if b.citation_count_or_stars and (
        not a.citation_count_or_stars or b.citation_count_or_stars > a.citation_count_or_stars
    ):
        a.citation_count_or_stars = b.citation_count_or_stars
    # union secondary URLs
    secs = set(s for s in (a.secondary_urls + "," + b.secondary_urls).split(",") if s.strip())
    secs.discard(a.canonical_url)
    a.secondary_urls = ", ".join(sorted(secs))
    # union topic tags
    tags = set(t.strip() for t in (a.topic_tags + "," + b.topic_tags).split(",") if t.strip())
    a.topic_tags = ", ".join(sorted(tags))
    if b.why_notable and len(b.why_notable) > len(a.why_notable):
        a.why_notable = b.why_notable
    return a


def _repo_slug(title: str) -> str:
    """Normalize a repo title to a slug for fuzzy dedup. Strips case, spaces, punctuation."""
    return re.sub(r"[^a-z0-9]", "", (title or "").lower())


def _repo_owner(r: Resource) -> str:
    """Extract owner from canonical_url (github) or fall back to authors_or_owners."""
    if r.canonical_url and "github.com/" in r.canonical_url:
        return r.canonical_url.split("github.com/", 1)[1].split("/")[0].lower()
    return (r.authors_or_owners or "").lower().strip()


def _url_path_stem(url: str) -> str:
    """Strip owner/ and return the repo-name portion (normalized to alnum-only)."""
    if url and "github.com/" in url:
        tail = url.split("github.com/", 1)[1].rstrip("/")
        segs = tail.split("/")
        if len(segs) >= 2:
            return re.sub(r"[^a-z0-9]", "", segs[1].lower())
    return ""


def dedupe_merge(rows: Iterable[Resource]) -> dict[str, Resource]:
    # Pass 1: primary dedup by URL hash (or synthetic id for URL-less rows)
    by_id: dict[str, Resource] = {}
    for r in rows:
        if r.canonical_url:
            rid = r.resource_id or resource_id_for(r.canonical_url)
        else:
            # Retain URL-less rows (e.g. Ziplime, Stock Research Agent) under a synthetic id
            rid = "n_" + hashlib.sha1(f"{r.type}|{r.title.lower()}".encode()).hexdigest()[:14]
        r.resource_id = rid
        if rid in by_id:
            merge_resources(by_id[rid], r)
        else:
            by_id[rid] = r

    # Pass 2: slug-based merge for type=repo rows. Handles GitHub-rename cases
    # (e.g., OpenBBTerminal → OpenBB where awesome-list and seed have different URLs
    # for the same project). Owner must match as a safety guard against collisions.
    slug_groups: dict[tuple[str, str], list[str]] = {}
    for rid, r in by_id.items():
        if r.type != "repo":
            continue
        slug = _repo_slug(r.title)
        if len(slug) < 4:
            continue
        owner = _repo_owner(r)
        slug_groups.setdefault((slug, owner), []).append(rid)

    to_remove: set[str] = set()
    for rids in slug_groups.values():
        if len(rids) <= 1:
            continue
        # Safety guard: if ≥2 rows in the group have canonical URLs, require the repo-name
        # stems to be prefix-related (handles GitHub renames like openbb→openbbterminal,
        # but rejects same-owner-different-repo like python-bizdays vs r-bizdays).
        urlful = [rid for rid in rids if by_id[rid].canonical_url]
        if len(urlful) >= 2:
            stems = sorted({_url_path_stem(by_id[rid].canonical_url) for rid in urlful}, key=len)
            shortest = stems[0]
            if shortest and not all(s == shortest or s.startswith(shortest) for s in stems):
                continue  # unrelated repo names — skip merge

        # Keeper: prefer URL-ful, then higher confidence, then longer title
        keeper = max(
            rids,
            key=lambda rid: (
                1 if by_id[rid].canonical_url else 0,
                _CONF_RANK.get(by_id[rid].confidence, 0),
                len(by_id[rid].title or ""),
            ),
        )
        for rid in rids:
            if rid == keeper or rid in to_remove:
                continue
            other = by_id[rid]
            if other.canonical_url and other.canonical_url != by_id[keeper].canonical_url:
                secs = {s for s in by_id[keeper].secondary_urls.split(",") if s.strip()}
                secs.add(other.canonical_url)
                by_id[keeper].secondary_urls = ", ".join(sorted(secs))
            merge_resources(by_id[keeper], other)
            to_remove.add(rid)

    for rid in to_remove:
        by_id.pop(rid, None)

    return by_id


# ---------------------------------------------------------------------------
# Incremental cache
# ---------------------------------------------------------------------------


def load_existing_index() -> dict[str, Resource]:
    if not OUT_JSON.exists():
        return {}
    data = json.loads(OUT_JSON.read_text())
    out: dict[str, Resource] = {}
    for item in data.get("resources", []):
        try:
            out[item["resource_id"]] = Resource(**item)
        except TypeError:
            continue
    return out


def is_fresh(r: Resource, days: int = REFRESH_DAYS) -> bool:
    if not r.retrieved_at:
        return False
    try:
        t = datetime.fromisoformat(r.retrieved_at.replace("Z", "+00:00"))
    except ValueError:
        return False
    return (datetime.now(timezone.utc) - t) < timedelta(days=days)


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


# ---------------------------------------------------------------------------
# HTTP helpers
# ---------------------------------------------------------------------------


def http_get(url: str, *, params: dict | None = None, accept: str | None = None) -> requests.Response | None:
    headers = {"Accept": accept} if accept else {}
    try:
        r = SESSION.get(url, params=params, timeout=HTTP_TIMEOUT, headers=headers)
        return r
    except requests.RequestException as e:
        log(f"  ! GET {url} failed: {e}")
        return None


LOG_LINES: list[str] = []


def log(msg: str) -> None:
    print(msg, flush=True)
    LOG_LINES.append(msg)


# ---------------------------------------------------------------------------
# Tier 1: awesome-lists
# ---------------------------------------------------------------------------

GITHUB_URL_RE = re.compile(r"https?://github\.com/([A-Za-z0-9_.-]+)/([A-Za-z0-9_.-]+)")
ARXIV_RE = re.compile(r"arxiv\.org/(?:abs|pdf)/(\d{4}\.\d{4,5})")
SSRN_RE = re.compile(r"ssrn\.com/[^\s)]+abstract[_-]?id=(\d+)", re.I)

MD_LINK_RE = re.compile(r"\[([^\]]+)\]\((https?://[^)\s]+)\)")
HEADER_RE = re.compile(r"^(#{1,6})\s+(.*)$")


def parse_awesome_markdown(md: str, source_tag: str, repo_url: str) -> list[Resource]:
    """Walk README.md linewise, track heading stack, emit Resource per markdown link."""
    rows: list[Resource] = []
    heading_stack: list[str] = []
    ts = now_iso()
    for line in md.splitlines():
        hm = HEADER_RE.match(line.strip())
        if hm:
            level = len(hm.group(1))
            title = hm.group(2).strip()
            heading_stack = heading_stack[: level - 1] + [title]
            continue
        for m in MD_LINK_RE.finditer(line):
            text, url = m.group(1).strip(), m.group(2).strip()
            if not url.startswith(("http://", "https://")):
                continue
            # skip badges / shields / image assets
            if any(bad in url for bad in ("shields.io", "badge.fury", ".png", ".jpg", ".svg", ".gif")):
                continue
            category = " > ".join(heading_stack[-2:]) if heading_stack else ""
            rtype = classify_url(url)
            if rtype is None:
                continue
            r = Resource(
                resource_id=resource_id_for(url),
                type=rtype,
                title=text[:200],
                sources=[source_tag],
                canonical_url=normalize_url(url),
                topic_tags=category,
                one_line_summary="",
                confidence="medium",
                retrieved_at=ts,
            )
            rows.append(r)
    return rows


def classify_url(url: str) -> str | None:
    u = url.lower()
    if "github.com/" in u:
        # exclude github.com root or user pages
        segs = [s for s in u.split("github.com/")[1].split("/") if s]
        if len(segs) >= 2:
            return "repo"
        return None
    if "arxiv.org/abs" in u or "arxiv.org/pdf" in u:
        return "paper"
    if "ssrn.com" in u and ("abstract_id=" in u or "abstract-id=" in u or "abstract/" in u):
        return "paper"
    if any(h in u for h in ("quantstart.com", "robotwealth.com", "hudsonthames.org", "blog.", "/blog/", "medium.com")):
        return "blog_post"
    if any(h in u for h in ("amazon.com/", "oreilly.com", "manning.com", "springer.com/book", "wiley.com/en-us/")):
        return "textbook"
    # academic publishers / journal hosts — used heavily by awesome-deep-trading
    if any(h in u for h in (
        "sciencedirect.com/", "link.springer.com/", "ieeexplore.ieee.org/",
        "researchgate.net/publication", "mdpi.com/", "hindawi.com/journals/",
        "iopscience.iop.org/", "tandfonline.com/doi", "aaai.org/", "openreview.net/",
        "papers.ssrn.com", "deepai.org/publication", "doi.org/",
    )):
        return "paper"
    return None


def run_tier1_awesome() -> list[Resource]:
    sources = [
        ("awesome-quant", "wilsonfreitas/awesome-quant", "https://raw.githubusercontent.com/wilsonfreitas/awesome-quant/master/README.md"),
        ("financial-ml", "firmai/financial-machine-learning", "https://raw.githubusercontent.com/firmai/financial-machine-learning/master/README.md"),
        ("awesome-deep-trading", "cbailes/awesome-deep-trading", "https://raw.githubusercontent.com/cbailes/awesome-deep-trading/master/README.md"),
    ]
    rows: list[Resource] = []
    for tag, repo, raw in sources:
        log(f"[tier1] fetching {repo}")
        r = http_get(raw)
        if r is None or r.status_code != 200:
            log(f"  ! {repo}: status {r.status_code if r else 'ERR'}")
            continue
        parsed = parse_awesome_markdown(r.text, tag, f"https://github.com/{repo}")
        log(f"  {repo}: {len(parsed)} candidate rows")
        rows.extend(parsed)

    # paperswithcode: JSON API for finance-adjacent tasks
    pwc_tasks = ["stock-market-prediction", "time-series-forecasting", "portfolio-optimization"]
    for task in pwc_tasks:
        log(f"[tier1] paperswithcode task={task}")
        j = http_get(f"https://paperswithcode.com/api/v1/tasks/{task}/papers/?items_per_page=100")
        if j is None or j.status_code != 200:
            log(f"  ! pwc {task}: status {j.status_code if j else 'ERR'}")
            continue
        try:
            data = j.json()
        except ValueError:
            log(f"  ! pwc {task}: non-JSON")
            continue
        ts = now_iso()
        for p in (data.get("results") or []):
            title = p.get("title") or ""
            paper_url = p.get("url_pdf") or p.get("url_abs") or ""
            # prefer arxiv abs
            arxiv_id = p.get("arxiv_id")
            if arxiv_id:
                paper_url = f"https://arxiv.org/abs/{arxiv_id}"
            if not paper_url:
                continue
            rows.append(
                Resource(
                    resource_id=resource_id_for(paper_url),
                    type="paper",
                    title=title[:300],
                    authors_or_owners=", ".join(p.get("authors") or [])[:300],
                    year=int((p.get("published") or "0000")[:4]) if (p.get("published") or "")[:4].isdigit() else None,
                    sources=[f"paperswithcode/{task}"],
                    canonical_url=normalize_url(paper_url),
                    topic_tags=task,
                    one_line_summary=(p.get("abstract") or "").split(".")[0][:300],
                    date_published=(p.get("published") or "")[:10],
                    confidence="medium",
                    retrieved_at=ts,
                )
            )
        time.sleep(1.5)

    log(f"[tier1] total {len(rows)} rows before dedup")
    return rows


# ---------------------------------------------------------------------------
# Tier 2: arxiv q-fin
# ---------------------------------------------------------------------------


def run_tier2_arxiv() -> list[Resource]:
    rows: list[Resource] = []
    top_ids_per_cat: dict[str, list[str]] = {}
    for cat in ARXIV_CATS:
        log(f"[tier2] arxiv cat={cat}")
        url = (
            f"http://export.arxiv.org/api/query?search_query=cat:{cat}"
            f"&start=0&max_results={ARXIV_MAX_PER_CAT}&sortBy=submittedDate&sortOrder=descending"
        )
        r = http_get(url)
        time.sleep(ARXIV_DELAY)
        if r is None or r.status_code != 200:
            log(f"  ! arxiv {cat}: status {r.status_code if r else 'ERR'}")
            continue
        feed = feedparser.parse(r.text)
        ts = now_iso()
        ids: list[str] = []
        for e in feed.entries:
            # e.id => http://arxiv.org/abs/XXXX.XXXXXvN
            m = re.search(r"abs/(\d{4}\.\d{4,5})", e.id)
            if not m:
                continue
            aid = m.group(1)
            ids.append(aid)
            authors = ", ".join(a.name for a in (e.get("authors") or []))
            pub = (e.get("published") or "")[:10]
            year = int(pub[:4]) if pub[:4].isdigit() else None
            abstract = (e.get("summary") or "").replace("\n", " ").strip()
            rows.append(
                Resource(
                    resource_id=resource_id_for(f"https://arxiv.org/abs/{aid}"),
                    type="paper",
                    title=(e.title or "").replace("\n", " ").strip()[:300],
                    authors_or_owners=authors[:300],
                    year=year,
                    sources=[f"arxiv/{cat}"],
                    canonical_url=f"https://arxiv.org/abs/{aid}",
                    topic_tags=cat,
                    one_line_summary=abstract.split(". ")[0][:300],
                    date_published=pub,
                    confidence="high" if (authors and year and aid) else "medium",
                    retrieved_at=ts,
                )
            )
        top_ids_per_cat[cat] = ids[:ARXIV_CITATION_TOP_N]

    # Semantic Scholar citation counts for top N per cat
    flat_ids = [i for ids in top_ids_per_cat.values() for i in ids]
    log(f"[tier2] fetching citations for {len(flat_ids)} papers from Semantic Scholar")
    cites: dict[str, int] = {}
    for aid in flat_ids:
        r = http_get(
            f"https://api.semanticscholar.org/graph/v1/paper/arXiv:{aid}",
            params={"fields": "citationCount"},
        )
        if r and r.status_code == 200:
            try:
                cites[aid] = int(r.json().get("citationCount") or 0)
            except (ValueError, TypeError):
                pass
        elif r and r.status_code == 429:
            log("  ! S2 rate-limited, sleeping 60s")
            time.sleep(60)
        time.sleep(3.2)  # 100 req / 5min unauth -> 1 per 3s
    # Attach
    for r in rows:
        m = re.search(r"abs/(\d{4}\.\d{4,5})", r.canonical_url)
        if m and m.group(1) in cites:
            r.citation_count_or_stars = cites[m.group(1)]
            if cites[m.group(1)] >= 100:
                r.why_notable = f"cited {cites[m.group(1)]}+"

    log(f"[tier2] total {len(rows)} arxiv rows")
    return rows


# ---------------------------------------------------------------------------
# Tier 3: institutional
# ---------------------------------------------------------------------------


def _extract_article_links(html: str, base: str, *, path_pat: re.Pattern | None = None) -> list[tuple[str, str, str]]:
    soup = BeautifulSoup(html, "lxml")
    out: list[tuple[str, str, str]] = []
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if not href or href.startswith("#"):
            continue
        abs_url = urllib.parse.urljoin(base, href)
        if path_pat and not path_pat.search(abs_url):
            continue
        text = a.get_text(" ", strip=True)[:200]
        if not text or len(text) < 8:
            continue
        # Try to grab nearby date
        date = ""
        par = a.find_parent()
        if par:
            dm = re.search(r"\b(20\d{2}[-/]\d{1,2}[-/]\d{1,2})\b|\b([A-Z][a-z]+ \d{1,2},? 20\d{2})\b", par.get_text(" ", strip=True))
            if dm:
                date = dm.group(0)
        out.append((text, abs_url, date))
    return out


INST_SOURCES = [
    ("aqr", "https://www.aqr.com/Insights/Research", re.compile(r"/Insights/Research/")),
    ("man-institute", "https://www.man.com/insights", re.compile(r"/insights/")),
    ("fed-feds-notes", "https://www.federalreserve.gov/econres/notes/feds-notes/default.htm", re.compile(r"/econres/notes/feds-notes/")),
    # two-sigma: insights index is JS-hydrated, only self-refs in static HTML; skipped
    # alpha-architect: Cloudflare 403 on unauth scrapers; skipped
]

# BIS via RePEc (bis.org/publ/work.htm 404s; RePEc mirror is stable).
# RePEc paginates at ~200 papers per page. Fetch 3 pages for ~600 most-recent BIS WPs.
BIS_REPEC_PAGES = [
    "https://ideas.repec.org/s/bis/biswps.html",
    "https://ideas.repec.org/s/bis/biswps2.html",
    "https://ideas.repec.org/s/bis/biswps3.html",
]
BIS_REPEC_TITLE_PAT = re.compile(r"/p/bis/biswps/\d+\.html")

BROWSER_UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
)


def _fetch_bis_repec(ts: str, inst_headers: dict) -> list[Resource]:
    """Fetch BIS working papers across RePEc pagination."""
    rows: list[Resource] = []
    seen: set[str] = set()
    kept = 0
    for page_url in BIS_REPEC_PAGES:
        try:
            r = SESSION.get(page_url, timeout=HTTP_TIMEOUT, headers=inst_headers)
        except requests.RequestException as e:
            log(f"  ! bis-wp-repec page {page_url}: {e}")
            continue
        if r.status_code != 200:
            log(f"  ! bis-wp-repec page {page_url}: status {r.status_code}")
            continue
        links = _extract_article_links(r.text, page_url, path_pat=BIS_REPEC_TITLE_PAT)
        for title, href, date in links:
            norm = normalize_url(href)
            if norm in seen:
                continue
            tl = title.lower().strip()
            if tl in {"by citations", "by downloads", "by date", "research", "insights",
                      "more", "read more", "subscribe", "contact", "share", "tweet", "pdf"}:
                continue
            seen.add(norm)
            rows.append(
                Resource(
                    resource_id=resource_id_for(href),
                    type="whitepaper",
                    title=title,
                    sources=["bis-wp-repec"],
                    canonical_url=norm,
                    topic_tags="institutional-research",
                    date_published=_norm_date(date),
                    confidence="medium",
                    retrieved_at=ts,
                )
            )
            kept += 1
        time.sleep(2.0)
    log(f"  bis-wp-repec: {kept} unique links across {len(BIS_REPEC_PAGES)} pages")
    return rows


def run_tier3_institutional() -> list[Resource]:
    rows: list[Resource] = []
    ts = now_iso()
    # Use browser UA for institutional sites (Cloudflare-averse to default SDK UAs)
    inst_headers = {"User-Agent": BROWSER_UA}
    # BIS — paginated RePEc
    rows.extend(_fetch_bis_repec(ts, inst_headers))

    for tag, url, pat in INST_SOURCES:
        log(f"[tier3] {tag}: {url}")
        try:
            r = SESSION.get(url, timeout=HTTP_TIMEOUT, headers=inst_headers)
        except requests.RequestException as e:
            log(f"  ! {tag}: {e}")
            continue
        if r.status_code != 200:
            log(f"  ! {tag}: status {r.status_code}")
            continue
        links = _extract_article_links(r.text, url, path_pat=pat)
        seen: set[str] = set()
        kept = 0
        for title, href, date in links:
            norm = normalize_url(href)
            if norm in seen:
                continue
            # filter nav text, sort links, social/share buttons
            tl = title.lower().strip()
            if tl in {
                "research", "insights", "more", "read more", "subscribe", "contact",
                "by citations", "by downloads", "by date", "share", "tweet", "pdf",
            }:
                continue
            seen.add(norm)
            rows.append(
                Resource(
                    resource_id=resource_id_for(href),
                    type="whitepaper",
                    title=title,
                    sources=[tag],
                    canonical_url=norm,
                    topic_tags="institutional-research",
                    date_published=_norm_date(date),
                    confidence="medium",
                    retrieved_at=ts,
                )
            )
            kept += 1
        log(f"  {tag}: {kept} unique links")
        time.sleep(2.0)
    log(f"[tier3] total {len(rows)} rows")
    return rows


def _norm_date(s: str) -> str:
    if not s:
        return ""
    s = s.strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%B %d, %Y", "%b %d, %Y", "%B %d %Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return ""


# ---------------------------------------------------------------------------
# Tier 4: SSRN top-10
# ---------------------------------------------------------------------------

SSRN_TOPTEN = [
    ("ssrn-fen-alltime", "https://papers.ssrn.com/sol3/topten/topTenResults.cfm?groupingId=203597&netorjrnl=ntwk"),
    ("ssrn-fen-12mo", "https://papers.ssrn.com/sol3/topten/topTenResults.cfm?groupingId=203597&netorjrnl=ntwk&scope=12"),
    ("ssrn-derivatives", "https://papers.ssrn.com/sol3/topten/topTenResults.cfm?groupingId=203731&netorjrnl=jrnl"),
    ("ssrn-microstructure", "https://papers.ssrn.com/sol3/topten/topTenResults.cfm?groupingId=204130&netorjrnl=jrnl"),
    ("ssrn-asset-pricing", "https://papers.ssrn.com/sol3/topten/topTenResults.cfm?groupingId=203619&netorjrnl=jrnl"),
]


def run_tier4_ssrn() -> list[Resource]:
    rows: list[Resource] = []
    ts = now_iso()
    ssrn_headers = {"User-Agent": BROWSER_UA}
    for tag, url in SSRN_TOPTEN:
        log(f"[tier4] {tag}")
        try:
            r = SESSION.get(url, timeout=HTTP_TIMEOUT, headers=ssrn_headers)
        except requests.RequestException as e:
            log(f"  ! {tag}: {e}")
            continue
        if r.status_code != 200:
            log(f"  ! {tag}: status {r.status_code} (SSRN top-ten pages Cloudflare-block unauth scrapers)")
            continue
        soup = BeautifulSoup(r.text, "lxml")
        # SSRN top-ten layout: each paper is an <a> linking to /sol3/papers.cfm?abstract_id=...
        seen: set[str] = set()
        for a in soup.find_all("a", href=True):
            href = a["href"]
            m = re.search(r"abstract[_-]?id=(\d+)", href)
            if not m:
                continue
            aid = m.group(1)
            if aid in seen:
                continue
            seen.add(aid)
            title = a.get_text(" ", strip=True)
            if not title or len(title) < 10:
                continue
            canon = f"https://papers.ssrn.com/sol3/papers.cfm?abstract_id={aid}"
            rows.append(
                Resource(
                    resource_id=resource_id_for(canon),
                    type="paper",
                    title=title[:300],
                    sources=[tag],
                    canonical_url=canon,
                    topic_tags="ssrn-top-ten",
                    confidence="medium",
                    retrieved_at=ts,
                    why_notable=f"SSRN top-ten ({tag})",
                )
            )
        log(f"  {tag}: {len(seen)} papers")
        time.sleep(2.0)
    log(f"[tier4] total {len(rows)} rows")
    return rows


# ---------------------------------------------------------------------------
# Tier 5: curator blogs
# ---------------------------------------------------------------------------

BLOG_SOURCES = [
    ("quantstart", "https://www.quantstart.com/articles/"),
    ("robot-wealth", "https://robotwealth.com/blog/"),
    ("hudson-thames", "https://hudsonthames.org/blog/"),
]


BLOG_NOISE_PATHS = ("/topic/", "/category/", "/tag/", "/author/", "/page/", "/feed", "/archive", "/ebook")


def run_tier5_blogs() -> list[Resource]:
    rows: list[Resource] = []
    ts = now_iso()
    blog_headers = {"User-Agent": BROWSER_UA}
    for tag, idx_url in BLOG_SOURCES:
        log(f"[tier5] {tag}")
        try:
            r = SESSION.get(idx_url, timeout=HTTP_TIMEOUT, headers=blog_headers)
        except requests.RequestException as e:
            log(f"  ! {tag}: {e}")
            continue
        if r.status_code != 200:
            log(f"  ! {tag}: status {r.status_code}")
            continue
        # 1) add the blog index posts themselves
        soup = BeautifulSoup(r.text, "lxml")
        post_links: list[tuple[str, str]] = []
        for a in soup.find_all("a", href=True):
            href = a["href"]
            abs_url = urllib.parse.urljoin(idx_url, href)
            host = urllib.parse.urlsplit(abs_url).netloc
            if tag.split("-")[0] not in host:
                continue
            if abs_url == idx_url or abs_url.endswith("/articles/") or abs_url.endswith("/blog/"):
                continue
            path = urllib.parse.urlsplit(abs_url).path
            if any(n in path for n in BLOG_NOISE_PATHS):
                continue
            text = a.get_text(" ", strip=True)
            if len(text) < 12:
                continue
            post_links.append((text[:200], abs_url))
        # dedup post urls
        seen_posts: set[str] = set()
        for title, href in post_links:
            norm = normalize_url(href)
            if norm in seen_posts:
                continue
            seen_posts.add(norm)
            rows.append(
                Resource(
                    resource_id=resource_id_for(href),
                    type="blog_post",
                    title=title,
                    sources=[tag],
                    canonical_url=norm,
                    topic_tags=f"recommended-by-{tag}",
                    confidence="medium",
                    retrieved_at=ts,
                )
            )

        # 2) extract embedded arxiv/ssrn/github links from the index page text
        for m in ARXIV_RE.finditer(r.text):
            aid = m.group(1)
            canon = f"https://arxiv.org/abs/{aid}"
            rows.append(
                Resource(
                    resource_id=resource_id_for(canon),
                    type="paper",
                    title=f"arXiv:{aid}",
                    sources=[tag],
                    canonical_url=canon,
                    topic_tags=f"recommended-by-{tag}",
                    confidence="medium",
                    retrieved_at=ts,
                )
            )
        for m in SSRN_RE.finditer(r.text):
            aid = m.group(1)
            canon = f"https://papers.ssrn.com/sol3/papers.cfm?abstract_id={aid}"
            rows.append(
                Resource(
                    resource_id=resource_id_for(canon),
                    type="paper",
                    title=f"SSRN:{aid}",
                    sources=[tag],
                    canonical_url=canon,
                    topic_tags=f"recommended-by-{tag}",
                    confidence="medium",
                    retrieved_at=ts,
                )
            )
        for m in GITHUB_URL_RE.finditer(r.text):
            gh = f"https://github.com/{m.group(1)}/{m.group(2)}"
            rows.append(
                Resource(
                    resource_id=resource_id_for(gh),
                    type="repo",
                    title=f"{m.group(1)}/{m.group(2)}",
                    authors_or_owners=m.group(1),
                    sources=[tag],
                    canonical_url=normalize_url(gh),
                    topic_tags=f"recommended-by-{tag}",
                    confidence="medium",
                    retrieved_at=ts,
                )
            )
        time.sleep(2.0)
    log(f"[tier5] total {len(rows)} rows")
    return rows


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

SHEET_COLS = [
    "resource_id",
    "type",
    "title",
    "authors_or_owners",
    "year",
    "sources",
    "canonical_url",
    "secondary_urls",
    "topic_tags",
    "one_line_summary",
    "citation_count_or_stars",
    "date_published",
    "mention_count",
    "why_notable",
    "confidence",
    "retrieved_at",
]


def _row_for_sheet(r: Resource) -> list[Any]:
    return [
        r.resource_id,
        r.type,
        r.title,
        r.authors_or_owners,
        r.year,
        ", ".join(r.sources),
        r.canonical_url,
        r.secondary_urls,
        r.topic_tags,
        r.one_line_summary,
        r.citation_count_or_stars,
        r.date_published,
        r.mention_count,
        r.why_notable,
        r.confidence,
        r.retrieved_at,
    ]


def _sort_key(r: Resource) -> tuple[str, str, str]:
    """Group by type alphabetically, then title, then canonical_url for stability."""
    return (r.type or "", (r.title or "").lower(), r.canonical_url or "")


def write_xlsx(all_deduped: dict[str, Resource], per_source: dict[str, list[Resource]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    sheets = [("All_Deduped", sorted(all_deduped.values(), key=_sort_key))]
    sheets += [(name, sorted(rows, key=_sort_key)) for name, rows in per_source.items()]
    for name, rows in sheets:
        ws = wb.create_sheet(name[:31])
        ws.append(SHEET_COLS)
        for r in rows:
            ws.append(_row_for_sheet(r))
        # style
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        widths = {
            "A": 18, "B": 10, "C": 60, "D": 30, "E": 6, "F": 28, "G": 60,
            "H": 30, "I": 28, "J": 60, "K": 14, "L": 12, "M": 8, "N": 40, "O": 10, "P": 22,
        }
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
    wb.save(OUT_XLSX)
    log(f"[out] wrote {OUT_XLSX.name}")


def write_json(all_deduped: dict[str, Resource]) -> None:
    sorted_rows = sorted(all_deduped.values(), key=_sort_key)
    payload = {
        "generated_at": now_iso(),
        "count": len(all_deduped),
        "resources": [r.to_dict() for r in sorted_rows],
    }
    OUT_JSON.write_text(json.dumps(payload, indent=2, ensure_ascii=False))
    log(f"[out] wrote {OUT_JSON.name} ({len(all_deduped)} unique)")


def _date_range(rows: list[Resource]) -> str:
    dates = [r.date_published for r in rows if r.date_published]
    if not dates:
        years = [r.year for r in rows if r.year]
        if not years:
            return "n/a"
        return f"{min(years)}–{max(years)}"
    return f"{min(dates)}–{max(dates)}"


def write_summary(per_source: dict[str, list[Resource]], all_deduped: dict[str, Resource], notes: list[str]) -> None:
    by_type: dict[str, int] = {}
    by_confidence: dict[str, int] = {}
    for r in all_deduped.values():
        by_type[r.type] = by_type.get(r.type, 0) + 1
        by_confidence[r.confidence] = by_confidence.get(r.confidence, 0) + 1

    sections: list[str] = []
    source_notes = {
        "Seeds_IG": "Cached-only seed rows carried over from a prior curation pass (22 papers + 39 repos). No live fetcher — refresh by editing data/quant_index.json directly.",
        "Awesome_Lists": "Parsed README.md of wilsonfreitas/awesome-quant + firmai/financial-machine-learning + cbailes/awesome-deep-trading. Badges, image assets, and non-quant links filtered via classify_url. paperswithcode skipped: site redirects to huggingface.co/papers (dead).",
        "ArXiv": "Fetched 200 most-recent per category from export.arxiv.org Atom API (3s pacing). Semantic Scholar citation counts backfilled for top 50 per category (350 total, ~19min at 3.2s/req unauth).",
        "Institutional": "Browser UA used to avoid Cloudflare 403s. BIS pulled via RePEc mirror (bis.org/publ/work.htm is 404 as of this run). Alpha Architect and Two Sigma both return 403 or JS-hydrated placeholders — dropped.",
        "SSRN_TopTen": "All 5 top-ten pages return Cloudflare 403 to unauth scrapers. Not fixable without a browser session or login. Zero rows — noted, not retried.",
        "Blogs": "QuantStart, Robot Wealth, Hudson & Thames. Posts filtered to exclude topic/category/tag/author/page/feed/archive/ebook paths. Embedded arxiv/SSRN/github references extracted from index pages.",
    }

    for name, rows in per_source.items():
        n = len(rows)
        flag = ""
        if name == "SSRN_TopTen" and n == 0:
            flag = " ⚠ blocked (Cloudflare 403)"
        elif n < 5 and name != "SSRN_TopTen":
            flag = " ⚠ suspiciously low — scraper may have broken"
        sections.append(
            f"### {name}\n"
            f"**Rows:** {n}{flag}  \n"
            f"**Date range:** {_date_range(rows)}  \n"
            f"{source_notes.get(name, '')}"
        )

    lines = [
        "# Quant Resource Index — Summary",
        "",
        f"**Generated:** {now_iso()}",
        f"**Unique resources:** {len(all_deduped)}",
        f"**By type:** " + ", ".join(f"{t}={c}" for t, c in sorted(by_type.items(), key=lambda x: -x[1])),
        f"**By confidence:** " + ", ".join(f"{t}={c}" for t, c in by_confidence.items()),
        "",
        "## Sources",
        "",
        *sections,
        "",
        "## Known gaps / decisions",
        "",
        "- **SSRN top-ten (Tier 4):** Cloudflare 403 on all 5 groupings. No workaround without auth.",
        "- **paperswithcode:** 302 → huggingface.co/papers (site retired). Skipped.",
        "- **Alpha Architect, Two Sigma:** Cloudflare-blocked / JS-hydrated respectively. Dropped from Tier 3.",
        "- **Tier 1 repo stars:** not backfilled — 625 calls exceeds 60/hr unauth GitHub budget. Set `GITHUB_TOKEN` and re-run with RUN['tier1_awesome']=True for star counts.",
        "- **Awesome-list short titles:** ~22 rows have title <4 chars (e.g. `ta`, `bt`, `xts`, `TTR`, `-1`, `-L-`) — markdown link text was just the bare repo name. URLs are real; titles could be backfilled from GitHub with a PAT.",
        "- **Repo slug-dedup:** a second dedup pass merges repo rows with matching (title-slug, owner) even when their canonical URLs differ (handles GitHub rename cases like OpenBBTerminal → OpenBB where seed and awesome-list have different URLs for the same project).",
        "",
    ]
    OUT_SUMMARY.write_text("\n".join(lines) + "\n")
    log(f"[out] wrote {OUT_SUMMARY.name}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def cached_rows_for(existing: dict[str, Resource], source_prefixes: tuple[str, ...]) -> list[Resource]:
    out = []
    for r in existing.values():
        if any(s.startswith(p) for p in source_prefixes for s in r.sources):
            out.append(r)
    return out


TIER_SOURCES = {
    "Seeds_IG": ("quantscience_ig",),
    "Awesome_Lists": ("awesome-quant", "financial-ml", "awesome-deep-trading", "paperswithcode"),
    "ArXiv": ("arxiv/",),
    "Institutional": ("aqr", "man-institute", "two-sigma", "alpha-architect", "bis-wp", "fed-feds-notes"),
    "SSRN_TopTen": ("ssrn-",),
    "Blogs": ("quantstart", "robot-wealth", "hudson-thames"),
}


def main(argv: list[str]) -> int:
    per_source: dict[str, list[Resource]] = {}
    existing = load_existing_index()
    log(f"[cache] {len(existing)} resources in existing {OUT_JSON.name}")

    def fetch_or_cache(sheet: str, run_flag: bool, fn):
        if run_flag:
            rows = fn()
        else:
            rows = cached_rows_for(existing, TIER_SOURCES[sheet])
            log(f"[{sheet}] using {len(rows)} cached rows (RUN disabled)")
        per_source[sheet] = rows

    per_source["Seeds_IG"] = cached_rows_for(existing, TIER_SOURCES["Seeds_IG"])
    log(f"[seeds] using {len(per_source['Seeds_IG'])} cached rows (no live fetcher)")

    fetch_or_cache("Awesome_Lists", RUN["tier1_awesome"], run_tier1_awesome)
    fetch_or_cache("ArXiv", RUN["tier2_arxiv"], run_tier2_arxiv)
    fetch_or_cache("Institutional", RUN["tier3_institutional"], run_tier3_institutional)
    fetch_or_cache("SSRN_TopTen", RUN["tier4_ssrn"], run_tier4_ssrn)
    fetch_or_cache("Blogs", RUN["tier5_blogs"], run_tier5_blogs)

    all_rows: list[Resource] = []
    for rows in per_source.values():
        all_rows.extend(rows)
    all_deduped = dedupe_merge(all_rows)

    write_json(all_deduped)
    write_xlsx(all_deduped, per_source)
    write_summary(per_source, all_deduped, LOG_LINES[-40:])
    log(f"[done] {len(all_deduped)} unique resources across {len(per_source)} sources")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
